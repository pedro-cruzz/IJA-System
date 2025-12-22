# ==========================
# IMPORTS PADRÃO PYTHON
# ==========================
import os
import re
import tempfile
import unicodedata
from datetime import date, datetime
from io import BytesIO
import json


from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
import uuid
import os

# ==========================
# FLASK
# ==========================
from flask import (Blueprint, after_this_request, current_app, flash, jsonify,
                   redirect, render_template, request, send_file,
                   send_from_directory, url_for)

from flask_login import current_user , login_required

# ==========================
# EXCEL / PDF
# ==========================
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import landscape
# ==========================
# SQLALCHEMY / BANCO
# ==========================
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import joinedload

# ==========================
# APP
# ==========================
from app import db
from app.models import Notificacao, Solicitacao, Usuario

print("--- ROTAS CARREGADAS COM SUCESSO ---")

bp = Blueprint('main', __name__)

@bp.context_processor
def inject_globals():
    notif_count = 0

    if current_user.is_authenticated:
        if current_user.tipo_usuario in ["admin", "operario", "visualizar"]:
            notif_count = (
                Notificacao.query
                .filter(
                    Notificacao.lida_em.is_(None),
                    Notificacao.apagada_em.is_(None),
                )
                .count()
            )
        else:
            notif_count = (
                Notificacao.query
                .filter(
                    Notificacao.usuario_id == current_user.id,
                    Notificacao.lida_em.is_(None),
                    Notificacao.apagada_em.is_(None),
                )
                .count()
            )

    return dict(notif_count=notif_count)



@bp.app_template_filter('datetimeformat')
def datetimeformat(value, format='%d-%m-%y'):
    try:
        # tenta converter string do tipo "2025-12-09"
        return datetime.strptime(value, "%Y-%m-%d").strftime(format)
    except:
        return value  # se falhar, retorna como está

def get_upload_folder():
    # pasta dentro do projeto: /seu_projeto/upload-files
    folder = os.path.join(current_app.root_path, '..', 'upload-files')
    folder = os.path.abspath(folder)
    os.makedirs(folder, exist_ok=True)
    return folder

ALLOWED_EXTENSIONS = {"pdf", "png", "jpg", "jpeg", "doc", "docx", "xls", "xlsx"}

def allowed_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

# --- DASHBOARD UVIS ---

@bp.route('/')
@login_required
def dashboard():

    # Se for admin, operario ou visualizar → painel admin
    if current_user.tipo_usuario in ['admin', 'operario', 'visualizar']:
        return redirect(url_for('main.admin_dashboard'))

    # Query base: solicitações SOMENTE do usuário logado
    query = Solicitacao.query.filter_by(usuario_id=current_user.id)

    # Filtro por status (?status=...)
    filtro_status = request.args.get('status')
    if filtro_status:
        query = query.filter(Solicitacao.status == filtro_status)

    # Paginação
    page = request.args.get("page", 1, type=int)
    paginacao = query.order_by(
        Solicitacao.data_criacao.desc()
    ).paginate(page=page, per_page=6, error_out=False)

    return render_template(
        'dashboard.html',
        solicitacoes=paginacao.items,
        paginacao=paginacao
    )

# --- PAINEL DE GESTÃO (Visualização para todos) ---
from flask_login import login_required, current_user
from datetime import datetime

@bp.route('/admin')
@login_required
def admin_dashboard():

    # 🔐 Controle de acesso
    if current_user.tipo_usuario not in ['admin', 'operario', 'visualizar']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('main.dashboard'))

    # Pode editar apenas admin e operario
    is_editable = current_user.tipo_usuario in ['admin', 'operario']

    # --- Captura filtros ---
    filtro_status = request.args.get("status")
    filtro_unidade = request.args.get("unidade")
    filtro_regiao = request.args.get("regiao")

    # --- Query base ---
    query = Solicitacao.query.join(Usuario)

    # --- Aplicação dos filtros ---
    if filtro_status:
        query = query.filter(Solicitacao.status == filtro_status)

    if filtro_unidade:
        query = query.filter(
            Usuario.nome_uvis.ilike(f"%{filtro_unidade}%")
        )

    if filtro_regiao:
        query = query.filter(
            Usuario.regiao.ilike(f"%{filtro_regiao}%")
        )

    # Paginação
    page = request.args.get("page", 1, type=int)

    paginacao = query.order_by(
        Solicitacao.data_criacao.desc()
    ).paginate(page=page, per_page=6, error_out=False)

    return render_template(
        'admin.html',
        pedidos=paginacao.items,
        paginacao=paginacao,
        is_editable=is_editable,
        now=datetime.now()
    )


@bp.route('/admin/exportar_excel')
@login_required
def exportar_excel():

    # 🔐 Permissão: somente admin e operario
    if current_user.tipo_usuario not in ['admin', 'operario']:
        flash('Permissão negada para exportar.', 'danger')
        return redirect(url_for('main.admin_dashboard'))

    try:
        filtro_status = request.args.get("status")
        filtro_unidade = request.args.get("unidade")
        filtro_regiao = request.args.get("regiao")

        # Evita Lazy Loading no Postgres
        query = (
            db.session.query(Solicitacao)
            .join(Usuario)
            .options(joinedload(Solicitacao.usuario))
        )

        if filtro_status:
            query = query.filter(Solicitacao.status == filtro_status)

        if filtro_unidade:
            query = query.filter(
                Usuario.nome_uvis.ilike(f"%{filtro_unidade}%")
            )

        if filtro_regiao:
            query = query.filter(
                Usuario.regiao.ilike(f"%{filtro_regiao}%")
            )

        pedidos = query.order_by(
            Solicitacao.data_criacao.desc()
        ).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório de Solicitações"

        headers = [
            "ID", "Unidade", "Região", "Data Agendada", "Hora",
            "Endereço Completo", "Latitude", "Longitude",
            "Foco", "Tipo Visita", "Altura", "Apoio CET?",
            "Observação", "Status", "Protocolo", "Justificativa"
        ]

        header_fill = PatternFill(
            start_color="1F4E78",
            end_color="1F4E78",
            fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for row_num, p in enumerate(pedidos, 2):
            uvis_nome = p.usuario.nome_uvis if p.usuario else "Não informado"
            uvis_regiao = p.usuario.regiao if p.usuario else "Não informado"

            endereco_completo = (
                f"{p.logradouro or ''}, {p.numero or ''} - "
                f"{p.bairro or ''} - "
                f"{(p.cidade or '')}/{(p.uf or '')} - {p.cep or ''}"
            )
            if p.complemento:
                endereco_completo += f" - {p.complemento}"

            data_formatada = ""
            if p.data_agendamento:
                if isinstance(p.data_agendamento, (date, datetime)):
                    data_formatada = p.data_agendamento.strftime("%d/%m/%Y")
                else:
                    data_formatada = str(p.data_agendamento)

            row = [
                p.id,
                uvis_nome,
                uvis_regiao,
                data_formatada,
                str(p.hora_agendamento or ""),
                endereco_completo,
                p.latitude or "",
                p.longitude or "",
                p.foco,
                p.tipo_visita or "",
                p.altura_voo or "",
                "SIM" if p.apoio_cet else "NÃO",
                p.observacao or "",
                p.status,
                p.protocolo or "",
                p.justificativa or ""
            ]

            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        ws.freeze_panes = "A2"

        for col in ws.columns:
            max_length = max(
                len(str(cell.value)) if cell.value else 0
                for cell in col
            )
            ws.column_dimensions[col[0].column_letter].width = min(
                max_length + 2, 50
            )

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            download_name="relatorio_solicitacoes.xlsx",
            as_attachment=False,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        db.session.rollback()
        print(f"ERRO EXPORTAR EXCEL: {e}")
        flash(
            "Erro ao gerar o Excel. Verifique se os dados estão corretos.",
            "danger"
        )
        return redirect(url_for('main.admin_dashboard'))

@bp.route('/admin/atualizar/<int:id>', methods=['POST'])
@login_required
def atualizar(id):

    # 🔐 Permissão
    if current_user.tipo_usuario not in ['admin', 'operario']:
        flash('Permissão negada para esta ação.', 'danger')
        return redirect(url_for('main.admin_dashboard'))

    pedido = Solicitacao.query.get_or_404(id)

    # --- Atualização de campos ---
    pedido.protocolo = request.form.get('protocolo')
    pedido.status = request.form.get('status')
    pedido.justificativa = request.form.get('justificativa')
    pedido.latitude = request.form.get('latitude')
    pedido.longitude = request.form.get('longitude')

    # --- Upload de anexo ---
    file = request.files.get("anexo")

    if file and file.filename:
        if not allowed_file(file.filename):
            flash("Tipo de arquivo não permitido.", "warning")
            return redirect(url_for('main.admin_dashboard'))

        original = secure_filename(file.filename)
        ext = original.rsplit(".", 1)[1].lower()
        unique_name = f"sol_{pedido.id}_{uuid.uuid4().hex}.{ext}"

        upload_folder = get_upload_folder()
        os.makedirs(upload_folder, exist_ok=True)

        save_path = os.path.join(upload_folder, unique_name)
        file.save(save_path)

        pedido.anexo_path = f"upload-files/{unique_name}"
        pedido.anexo_nome = original

    try:
        db.session.commit()
        flash('Pedido atualizado com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        print(f"ERRO ATUALIZAR PEDIDO: {e}")
        flash('Erro ao atualizar o pedido.', 'danger')

    return redirect(url_for('main.admin_dashboard'))


# --- NOVO PEDIDO ---
from flask_login import login_required, current_user

@bp.route('/novo_cadastro', methods=['GET', 'POST'], endpoint='novo')
@login_required
def novo():

    hoje = date.today().isoformat()

    if request.method == 'POST':
        try:
            # --- Data ---
            data_str = request.form.get('data')
            hora_str = request.form.get('hora')

            data_obj = (
                datetime.strptime(data_str, '%Y-%m-%d').date()
                if data_str else None
            )

            hora_obj = (
                datetime.strptime(hora_str, '%H:%M').time()
                if hora_str else None
            )

            apoio_cet_bool = request.form.get('apoio_cet') == 'sim'

            nova_solicitacao = Solicitacao(
                data_agendamento=data_obj,
                hora_agendamento=hora_obj,

                cep=request.form.get('cep'),
                logradouro=request.form.get('logradouro'),
                bairro=request.form.get('bairro'),
                cidade=request.form.get('cidade'),
                numero=request.form.get('numero'),
                uf=request.form.get('uf'),
                complemento=request.form.get('complemento'),

                foco=request.form.get('foco'),
                tipo_visita=request.form.get('tipo_visita'),
                altura_voo=request.form.get('altura_voo'),
                apoio_cet=apoio_cet_bool,
                observacao=request.form.get('observacao'),

                latitude=request.form.get('latitude'),
                longitude=request.form.get('longitude'),

                # 🔑 RELAÇÃO CORRETA COM FLASK-LOGIN
                usuario_id=current_user.id,

                status='PENDENTE'
            )

            db.session.add(nova_solicitacao)
            db.session.commit()

            flash('Pedido enviado com sucesso!', 'success')
            return redirect(url_for('main.dashboard'))

        except ValueError as ve:
            db.session.rollback()
            flash(f"Erro no formato de data ou hora.", "warning")

        except Exception as e:
            db.session.rollback()
            print(f"ERRO NOVO CADASTRO: {e}")
            flash("Erro ao salvar o pedido.", "danger")

    return render_template('cadastro.html', hoje=hoje)

# --- LOGIN ---
from flask_login import login_user

from flask_login import login_user, current_user

@bp.route('/login', methods=['GET', 'POST'])
def login():
    # Se já estiver logado, redireciona
    if current_user.is_authenticated:
        if current_user.tipo_usuario in ['admin', 'operario', 'visualizar']:
            return redirect(url_for('main.admin_dashboard'))
        return redirect(url_for('main.dashboard'))

    if request.method == 'POST':
        login_form = request.form.get('login')
        senha_form = request.form.get('senha')
        user = Usuario.query.filter_by(login=login_form).first()
        if user and user.check_senha(senha_form):
            login_user(user)  # 🔥 ÚNICO controle de login
            flash(
                f'Bem-vindo, {user.nome_uvis}! Login realizado com sucesso.',
                'success'
            )
            if user.tipo_usuario in ['admin', 'operario', 'visualizar']:
                return redirect(url_for('main.admin_dashboard'))
            return redirect(url_for('main.dashboard'))
        flash('Login ou senha incorretos. Tente novamente.', 'danger')

    return render_template('login.html')

# --- LOGOUT ---
from flask_login import logout_user, login_required

@bp.route('/logout')
@login_required
def logout():
    logout_user()          # 🔑 encerra o current_user
    session.clear()        # opcional (flash, tema, etc)
    flash('Você saiu do sistema.', 'info')
    return redirect(url_for('main.login'))


@bp.route("/forcar_erro")
def forcar_erro():
    1 / 0  # erro proposital
    return "nunca vai chegar aqui"

# Openpyxl (Excel)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
# ReportLab (PDF)
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (PageBreak, Paragraph, SimpleDocTemplate,
                                Spacer, Table, TableStyle)

# O objeto 'bp' precisa ser definido (Exemplo: bp = Blueprint('main', __name__))
# E 'Usuario' e 'Solicitacao' precisam ser seus modelos SQLAlchemy

# =======================================================================
# Função Auxiliar de Filtros (Reutilizada em todas as rotas)
# =======================================================================

def aplicar_filtros_base(query, filtro_data, uvis_id):
    """Aplica o filtro de mês/ano e opcionalmente o filtro de UVIS (usuario_id)."""
    # --- AJUSTE DE COMPATIBILIDADE PARA O RENDER (POSTGRES) ---
    if db.engine.name == 'postgresql':
        # No PostgreSQL usamos to_char
        query = query.filter(db.func.to_char(Solicitacao.data_criacao, 'YYYY-MM') == filtro_data)
    else:
        # No SQLite (seu PC) continuamos com strftime
        query = query.filter(db.func.strftime('%Y-%m', Solicitacao.data_criacao) == filtro_data)
    # Filtro de UVIS (opcional)
    if uvis_id:
        query = query.filter(Solicitacao.usuario_id == uvis_id)
        
    return query


from datetime import datetime

# =======================================================================
# ROTA 1: Visualização do Relatório (HTML)
# =======================================================================
from flask import redirect, render_template, request, session, url_for

from app import db
from app.models import Solicitacao, Usuario


@bp.route('/relatorios', methods=['GET'])
def relatorios():
    if not current_user.is_authenticated:
        return redirect(url_for('main.login'))

    try:
        # 🔹 Parâmetros de Filtro
        mes_atual = request.args.get('mes', datetime.now().month, type=int)
        ano_atual = request.args.get('ano', datetime.now().year, type=int)
        filtro_data = f"{ano_atual}-{mes_atual:02d}"

        # 🔐 Controle de UVIS
        if current_user.tipo_usuario == 'uvis':
            uvis_id = current_user.id
        else:
            uvis_id = request.args.get('uvis_id', type=int)

        # 🔹 UVIS disponíveis (admin / operário / visualizar)
        uvis_disponiveis = []
        if current_user.tipo_usuario in ['admin', 'operario', 'visualizar']:
            uvis_disponiveis = [
                (u.id, u.nome_uvis)
                for u in (
                    db.session.query(Usuario.id, Usuario.nome_uvis)
                    .filter(Usuario.tipo_usuario == 'uvis')
                    .order_by(Usuario.nome_uvis)
                    .all()
                )
            ]

        # 🔹 Histórico mensal (Postgres x SQLite)
        if db.engine.name == 'postgresql':
            func_mes = db.func.to_char(Solicitacao.data_criacao, 'YYYY-MM')
        else:
            func_mes = db.func.strftime('%Y-%m', Solicitacao.data_criacao)

        dados_mensais = [
            (mes, total)
            for mes, total in (
                db.session.query(func_mes.label('mes'), db.func.count(Solicitacao.id))
                .group_by('mes')
                .order_by('mes')
                .all()
            )
        ]

        anos_disponiveis = (
            sorted({m.split('-')[0] for m, _ in dados_mensais}, reverse=True)
            if dados_mensais else [ano_atual]
        )

        # 🔹 Query base única
        base_query = aplicar_filtros_base(
            db.session.query(Solicitacao),
            filtro_data,
            uvis_id
        )

        # 🔹 Totais
        total_solicitacoes = base_query.count()
        total_aprovadas = base_query.filter(Solicitacao.status == "APROVADO").count()
        total_aprovadas_com_recomendacoes = base_query.filter(
            Solicitacao.status == "APROVADO COM RECOMENDAÇÕES"
        ).count()
        total_recusadas = base_query.filter(Solicitacao.status == "NEGADO").count()
        total_analise = base_query.filter(Solicitacao.status == "EM ANÁLISE").count()
        total_pendentes = base_query.filter(Solicitacao.status == "PENDENTE").count()

        # 🔹 Agrupamentos (SEM Row)

        dados_regiao = [
            (regiao or "Não informado", total)
            for regiao, total in (
                aplicar_filtros_base(
                    db.session.query(Usuario.regiao, db.func.count(Solicitacao.id))
                    .join(Usuario),
                    filtro_data,
                    uvis_id
                )
                .group_by(Usuario.regiao)
                .all()
            )
        ]

        dados_status = [
            (status or "Não informado", total)
            for status, total in (
                base_query
                .with_entities(Solicitacao.status, db.func.count(Solicitacao.id))
                .group_by(Solicitacao.status)
                .all()
            )
        ]

        dados_foco = [
            (foco or "Não informado", total)
            for foco, total in (
                base_query
                .with_entities(Solicitacao.foco, db.func.count(Solicitacao.id))
                .group_by(Solicitacao.foco)
                .all()
            )
        ]

        dados_tipo_visita = [
            (tipo or "Não informado", total)
            for tipo, total in (
                base_query
                .with_entities(Solicitacao.tipo_visita, db.func.count(Solicitacao.id))
                .group_by(Solicitacao.tipo_visita)
                .all()
            )
        ]

        dados_altura_voo = [
            (altura or "Não informado", total)
            for altura, total in (
                base_query
                .with_entities(Solicitacao.altura_voo, db.func.count(Solicitacao.id))
                .group_by(Solicitacao.altura_voo)
                .all()
            )
        ]

        dados_unidade = [
            (uvis or "Não informado", total)
            for uvis, total in (
                aplicar_filtros_base(
                    db.session.query(Usuario.nome_uvis, db.func.count(Solicitacao.id))
                    .join(Usuario)
                    .filter(Usuario.tipo_usuario == 'uvis'),
                    filtro_data,
                    uvis_id
                )
                .group_by(Usuario.nome_uvis)
                .all()
            )
        ]

        return render_template(
            'relatorios.html',
            total_solicitacoes=total_solicitacoes,
            total_aprovadas=total_aprovadas,
            total_aprovadas_com_recomendacoes=total_aprovadas_com_recomendacoes,
            total_recusadas=total_recusadas,
            total_analise=total_analise,
            total_pendentes=total_pendentes,
            dados_regiao=dados_regiao,
            dados_status=dados_status,
            dados_foco=dados_foco,
            dados_tipo_visita=dados_tipo_visita,
            dados_altura_voo=dados_altura_voo,
            dados_unidade=dados_unidade,
            dados_mensais=dados_mensais,
            mes_selecionado=mes_atual,
            ano_selecionado=ano_atual,
            anos_disponiveis=anos_disponiveis,
            uvis_id_selecionado=uvis_id,
            uvis_disponiveis=uvis_disponiveis
        )

    except Exception as e:
        db.session.rollback()
        print(f"ERRO NOS RELATÓRIOS: {e}")
        return render_template(
            "erro.html",
            codigo=500,
            titulo="Erro nos Relatórios",
            mensagem="Houve um erro técnico ao processar os dados."
        )


import os
import tempfile
from datetime import datetime
from io import BytesIO

from flask import send_file, request
from flask_login import login_required, current_user
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    PageBreak, Image as RLImage
)

try:
    import matplotlib.pyplot as plt
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False

@bp.route('/admin/exportar_relatorio_pdf')
@login_required
def exportar_relatorio_pdf():
    # -------------------------
    # 1. Parâmetros e filtros (IGUAL ao /relatorios)
    # -------------------------
    mes = int(request.args.get('mes', datetime.now().month))
    ano = int(request.args.get('ano', datetime.now().year))
    orient = request.args.get('orient', default='portrait')  # 'portrait' ou 'landscape'
    filtro_data = f"{ano}-{mes:02d}"

    if current_user.tipo_usuario == 'uvis':
        uvis_id = current_user.id
    else:
        uvis_id = request.args.get('uvis_id', type=int)

    # -------------------------
    # 2. Query base e detalhe
    # -------------------------
    base_query = aplicar_filtros_base(
        db.session.query(Solicitacao),
        filtro_data,
        uvis_id
    )

    query_detalhe = aplicar_filtros_base(
        db.session.query(Solicitacao, Usuario).join(Usuario, Usuario.id == Solicitacao.usuario_id),
        filtro_data,
        uvis_id
    )

    query_results = query_detalhe.order_by(Solicitacao.data_criacao.desc()).all()

    # -------------------------
    # 3. Totais
    # -------------------------
    total_solicitacoes = base_query.count()
    total_aprovadas = base_query.filter(Solicitacao.status == "APROVADO").count()
    total_aprovadas_com_recomendacoes = base_query.filter(
        Solicitacao.status == "APROVADO COM RECOMENDAÇÕES"
    ).count()
    total_recusadas = base_query.filter(Solicitacao.status == "NEGADO").count()
    total_analise = base_query.filter(Solicitacao.status == "EM ANÁLISE").count()
    total_pendentes = base_query.filter(Solicitacao.status == "PENDENTE").count()

    STATUS_COLORS = {
        "APROVADO": "#2ecc71",
        "APROVADO COM RECOMENDAÇÕES": "#ee650a",
        "EM ANÁLISE": "#f1c40f",
        "PENDENTE": "#3498db",
        "NEGADO": "#e74c3c",
    }

    # -------------------------
    # 4. Agrupamentos
    # -------------------------
    dados_regiao = [
        (regiao or "Não informado", total)
        for regiao, total in (
            aplicar_filtros_base(
                db.session.query(Usuario.regiao, db.func.count(Solicitacao.id)).join(Usuario),
                filtro_data,
                uvis_id
            )
            .group_by(Usuario.regiao)
            .all()
        )
    ]

    dados_status = [
        (status or "Não informado", total)
        for status, total in (
            base_query
            .with_entities(Solicitacao.status, db.func.count(Solicitacao.id))
            .group_by(Solicitacao.status)
            .all()
        )
    ]

    dados_foco = [
        (foco or "Não informado", total)
        for foco, total in (
            base_query
            .with_entities(Solicitacao.foco, db.func.count(Solicitacao.id))
            .group_by(Solicitacao.foco)
            .all()
        )
    ]

    dados_tipo_visita = [
        (tipo or "Não informado", total)
        for tipo, total in (
            base_query
            .with_entities(Solicitacao.tipo_visita, db.func.count(Solicitacao.id))
            .group_by(Solicitacao.tipo_visita)
            .all()
        )
    ]

    dados_altura_voo = [
        (altura or "Não informado", total)
        for altura, total in (
            base_query
            .with_entities(Solicitacao.altura_voo, db.func.count(Solicitacao.id))
            .group_by(Solicitacao.altura_voo)
            .all()
        )
    ]

    dados_unidade = [
        (uvis_nome or "Não informado", total)
        for uvis_nome, total in (
            aplicar_filtros_base(
                db.session.query(Usuario.nome_uvis, db.func.count(Solicitacao.id))
                .join(Usuario)
                .filter(Usuario.tipo_usuario == 'uvis'),
                filtro_data,
                uvis_id
            )
            .group_by(Usuario.nome_uvis)
            .all()
        )
    ]

    if db.engine.name == 'postgresql':
        func_mes = db.func.to_char(Solicitacao.data_criacao, 'YYYY-MM')
    else:
        func_mes = db.func.strftime('%Y-%m', Solicitacao.data_criacao)

    dados_mensais = [
        tuple(row) for row in (
            db.session.query(func_mes.label('mes'), db.func.count(Solicitacao.id))
            .group_by('mes')
            .order_by('mes')
            .all()
        )
    ]

    # -------------------------
    # 5. Preparar PDF
    # -------------------------
    tmp_pdf = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    caminho_pdf = tmp_pdf.name
    tmp_pdf.close()

    pagesize = landscape(A4) if orient == 'landscape' else A4

    doc = SimpleDocTemplate(
        caminho_pdf,
        pagesize=pagesize,
        leftMargin=14*mm, rightMargin=14*mm,
        topMargin=16*mm, bottomMargin=16*mm
    )

    styles = getSampleStyleSheet()

    # Tipografia melhor
    title_style = ParagraphStyle(
        'title',
        parent=styles['Title'],
        fontSize=18,
        leading=22,
        alignment=1,
        textColor=colors.HexColor('#0d6efd'),
        spaceAfter=10
    )

    subtitle_style = ParagraphStyle(
        'subtitle',
        parent=styles['Normal'],
        fontSize=10,
        leading=14,
        alignment=1,
        textColor=colors.HexColor('#555'),
        spaceAfter=12
    )

    section_h = ParagraphStyle(
        'sec',
        parent=styles['Heading2'],
        fontSize=12,
        leading=16,
        textColor=colors.HexColor('#0d6efd'),
        spaceBefore=10,
        spaceAfter=6
    )

    normal = ParagraphStyle(
        'normal',
        parent=styles['Normal'],
        fontSize=9.5,
        leading=13
    )

    cell_style = ParagraphStyle(
        'cell',
        parent=styles['BodyText'],
        fontSize=8.6,
        leading=11,
        textColor=colors.HexColor('#222'),
        wordWrap='CJK',
        splitLongWords=True
    )

    story = []

    # -------------------------
    # CAPA (Resumo)
    # -------------------------
    story.append(Paragraph(f"Relatório Mensal — {mes:02d}/{ano}", title_style))

    filtro_txt = f"Filtro: {filtro_data}"
    if uvis_id:
        filtro_txt += f" | UVIS ID: {uvis_id}"
    else:
        filtro_txt += " | UVIS: Todas"
    story.append(Paragraph(filtro_txt, subtitle_style))

    # Cards do resumo (bem mais bonito)
    def resumo_cards():
        cards = [
            ("Total", total_solicitacoes, '#0d6efd'),
            ("Aprovadas", total_aprovadas, '#198754'),
            ("Aprov. c/ Recom.", total_aprovadas_com_recomendacoes, '#6c757d'),
            ("Negadas", total_recusadas, '#dc3545'),
            ("Em Análise", total_analise, '#ffc107'),
            ("Pendentes", total_pendentes, '#0dcaf0'),
        ]

        rows = []
        row = []
        for i, (label, value, hexcolor) in enumerate(cards, start=1):
            box = Table(
                [
                    [Paragraph(label, ParagraphStyle('l', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#666')))],
                    [Paragraph(str(value), ParagraphStyle('v', parent=styles['Normal'], fontSize=18, leading=20, textColor=colors.HexColor(hexcolor)))]
                ],
                colWidths=[48*mm] if orient == 'portrait' else [52*mm],
            )
            box.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#f8f9fa')),
                ('BOX', (0,0), (-1,-1), 0.6, colors.HexColor('#e5e7eb')),
                ('LEFTPADDING', (0,0), (-1,-1), 8),
                ('RIGHTPADDING', (0,0), (-1,-1), 8),
                ('TOPPADDING', (0,0), (-1,-1), 6),
                ('BOTTOMPADDING', (0,0), (-1,-1), 6),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ]))

            row.append(box)
            if len(row) == 3:
                rows.append(row)
                row = []

        if row:
            # completa a linha
            while len(row) < 3:
                row.append(Spacer(1, 1))
            rows.append(row)

        grid = Table(rows, colWidths=None)
        grid.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('LEFTPADDING', (0,0), (-1,-1), 0),
            ('RIGHTPADDING', (0,0), (-1,-1), 0),
            ('TOPPADDING', (0,0), (-1,-1), 0),
            ('BOTTOMPADDING', (0,0), (-1,-1), 0),
        ]))
        return grid

    story.append(resumo_cards())
    story.append(Spacer(1, 10))

    # -------------------------
    # TABELAS (DADOS ESCRITOS PRIMEIRO)
    # -------------------------
    def add_count_table(titulo, dados, col1="Categoria"):
        story.append(Paragraph(titulo, section_h))

        rows = [
            [Paragraph(col1, ParagraphStyle('th', parent=cell_style, textColor=colors.white, fontSize=9)),
             Paragraph("Total", ParagraphStyle('th2', parent=cell_style, textColor=colors.white, fontSize=9))]
        ]

        for nome, total in (dados or [("Nenhum", 0)]):
            rows.append([Paragraph(str(nome), cell_style), Paragraph(str(total), cell_style)])

        tbl = Table(rows, repeatRows=1, colWidths=[140*mm, 25*mm] if orient == 'portrait' else [190*mm, 30*mm])
        tbl.setStyle(TableStyle([
            ('BACKGROUND', (0,0),(-1,0),colors.HexColor('#0d6efd')),
            ('TEXTCOLOR',(0,0),(-1,0),colors.white),
            ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
            ('FONTSIZE',(0,0),(-1,0),9),
            ('GRID',(0,0),(-1,-1),0.25,colors.HexColor('#d9dee7')),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#fbfdff')]),
            ('VALIGN',(0,0),(-1,-1),'TOP'),
            ('LEFTPADDING',(0,0),(-1,-1),6),
            ('RIGHTPADDING',(0,0),(-1,-1),6),
            ('TOPPADDING',(0,0),(-1,-1),4),
            ('BOTTOMPADDING',(0,0),(-1,-1),4),
        ]))
        story.append(tbl)
        story.append(Spacer(1, 10))

    # Um “Resumo por agrupamento” em sequência (mais agradável)
    story.append(Paragraph("Resumo por Agrupamentos", section_h))
    story.append(Paragraph("Abaixo estão os agrupamentos do mês selecionado, apresentados em formato de tabela.", normal))
    story.append(Spacer(1, 6))

    add_count_table("Agrupamento — Região", dados_regiao)
    add_count_table("Agrupamento — Status", dados_status)
    add_count_table("Agrupamento — Foco", dados_foco)
    add_count_table("Agrupamento — Tipo de Visita", dados_tipo_visita)
    add_count_table("Agrupamento — Altura do Voo", dados_altura_voo)
    add_count_table("Agrupamento — Unidade (UVIS)", dados_unidade)
    add_count_table("Histórico Mensal (tabela)", dados_mensais, col1="Mês")

    # -------------------------
    # ✅ GRÁFICOS (AGORA DEPOIS DOS DADOS ESCRITOS)
    # -------------------------
    story.append(PageBreak())
    story.append(Paragraph("Gráficos", section_h))
    story.append(Paragraph("Os gráficos abaixo representam visualmente os dados apresentados nas tabelas anteriores.", normal))
    story.append(Spacer(1, 8))

    def safe_img_from_plt(fig, width_mm=170):
        bio = BytesIO()
        fig.tight_layout()
        fig.savefig(bio, format='png', dpi=220, bbox_inches='tight')
        plt.close(fig)
        bio.seek(0)
        return RLImage(bio, width=width_mm*mm)

    if MATPLOTLIB_AVAILABLE:
        try:
            # 1) Donut por status (mais limpo)
            labels = [s for s, _ in dados_status]
            values = [c for _, c in dados_status]
            colors_status = [STATUS_COLORS.get(s, "#bdc3c7") for s in labels]

            fig1, ax1 = plt.subplots(figsize=(6.4, 3.0))
            def autopct(p): return f'{p:.0f}%' if p >= 6 else ''
            wedges, *_ = ax1.pie(
                values or [1],
                labels=None,
                colors=colors_status,
                autopct=autopct,
                startangle=90,
                pctdistance=0.78,
                textprops={'fontsize': 9}
            )
            centre_circle = plt.Circle((0, 0), 0.58, fc='white')
            ax1.add_artist(centre_circle)
            ax1.legend(wedges, labels, loc='center left', bbox_to_anchor=(1.02, 0.5),
                       fontsize=9, frameon=False)
            ax1.set_title('Distribuição por Status', fontsize=11, pad=10)
            ax1.axis('equal')

            story.append(safe_img_from_plt(fig1, width_mm=170))
            story.append(Spacer(1, 10))

            # 2) Top UVIS (barra horizontal)
            u_names = [u for u, _ in dados_unidade[:10]]
            u_vals = [c for _, c in dados_unidade[:10]]

            fig2, ax2 = plt.subplots(figsize=(7.2, 3.0))
            ax2.barh(u_names[::-1] or ['Nenhum'], u_vals[::-1] or [0])
            ax2.set_xlabel('Total', fontsize=9)
            ax2.set_title('Top UVIS', fontsize=11, pad=10)
            ax2.tick_params(axis='both', labelsize=9)
            ax2.grid(axis='x', linestyle=':', linewidth=0.6, alpha=0.6)

            story.append(safe_img_from_plt(fig2, width_mm=180 if orient == 'landscape' else 170))
            story.append(Spacer(1, 10))

            # 3) Histórico mensal (linha)
            months = [m for m, _ in dados_mensais]
            counts = [c for _, c in dados_mensais]

            fig3, ax3 = plt.subplots(figsize=(7.2, 3.0))
            if months:
                ax3.plot(range(len(months)), counts, marker='o', linewidth=1.6)
                ax3.set_xticks(range(len(months)))
                ax3.set_xticklabels(months, rotation=45, ha='right', fontsize=9)
            ax3.set_title('Histórico Mensal', fontsize=11, pad=10)
            ax3.tick_params(axis='y', labelsize=9)
            ax3.grid(axis='y', linestyle=':', linewidth=0.6, alpha=0.6)

            story.append(safe_img_from_plt(fig3, width_mm=185 if orient == 'landscape' else 170))
            story.append(Spacer(1, 8))

        except Exception:
            story.append(Paragraph("Gráficos indisponíveis (erro ao gerar).", normal))
    else:
        story.append(Paragraph("Matplotlib não disponível — gráficos foram omitidos.", normal))

    # -------------------------
    # DETALHES (opcional: se quiser manter, deixa por último)
    # -------------------------
    story.append(PageBreak())
    story.append(Paragraph("Registros Detalhados", section_h))
    story.append(Paragraph("Listagem completa dos registros retornados pelo filtro selecionado.", normal))
    story.append(Spacer(1, 8))

    registros_header = [
        'Data', 'Hora', 'Unidade', 'Região', 'Protocolo',
        'Status', 'Foco', 'Tipo Visita', 'Altura Voo', 'Observação'
    ]
    registros_rows = [[Paragraph(h, ParagraphStyle('hdr', parent=cell_style, textColor=colors.white, fontSize=8.7))
                       for h in registros_header]]

    for s, u in query_results:
        data_str = s.data_criacao.strftime("%d/%m/%Y") if getattr(s, 'data_criacao', None) else ''
        hora_str = getattr(s, 'hora_agendamento', '')
        hora_str = hora_str.strftime("%H:%M") if hasattr(hora_str, 'strftime') else str(hora_str or '')

        unidade = getattr(u, 'nome_uvis', '') or "Não informado"
        regiao = getattr(u, 'regiao', '') or "Não informado"

        protocolo = getattr(s, 'protocolo', '') or ''
        status = getattr(s, 'status', '') or ''
        foco = getattr(s, 'foco', '') or ''
        tipo_visita = getattr(s, 'tipo_visita', '') or ''
        altura_voo = getattr(s, 'altura_voo', '') or ''
        obs = getattr(s, 'observacao', '') or ''

        registros_rows.append([
            Paragraph(str(data_str), cell_style),
            Paragraph(str(hora_str), cell_style),
            Paragraph(str(unidade), cell_style),
            Paragraph(str(regiao), cell_style),
            Paragraph(str(protocolo), cell_style),
            Paragraph(str(status), cell_style),
            Paragraph(str(foco), cell_style),
            Paragraph(str(tipo_visita), cell_style),
            Paragraph(str(altura_voo), cell_style),
            Paragraph(str(obs), cell_style),
        ])

    chunk_size = 26
    colWidths = [18*mm, 14*mm, 28*mm, 22*mm, 22*mm, 22*mm, 22*mm, 26*mm, 18*mm, 60*mm]

    for i in range(0, len(registros_rows), chunk_size):
        chunk = registros_rows[i:i+chunk_size]
        tbl = Table(chunk, repeatRows=1, colWidths=colWidths)
        tbl.setStyle(TableStyle([
            ('BACKGROUND', (0,0),(-1,0),colors.HexColor('#0d6efd')),
            ('TEXTCOLOR',(0,0),(-1,0),colors.white),
            ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
            ('FONTSIZE',(0,0),(-1,0),8.4),
            ('GRID',(0,0),(-1,-1),0.25,colors.HexColor('#d9dee7')),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#fbfdff')]),
            ('VALIGN',(0,0),(-1,-1),'TOP'),
            ('LEFTPADDING',(0,0),(-1,-1),4),
            ('RIGHTPADDING',(0,0),(-1,-1),4),
            ('TOPPADDING',(0,0),(-1,-1),3),
            ('BOTTOMPADDING',(0,0),(-1,-1),3),
        ]))
        story.append(tbl)
        story.append(Spacer(1, 6))
        if i + chunk_size < len(registros_rows):
            story.append(PageBreak())

    # -------------------------
    # Header/Footer
    # -------------------------
    def _header_footer(canvas, doc_):
        canvas.saveState()
        w, h = pagesize

        canvas.setFillColor(colors.HexColor('#0d6efd'))
        canvas.rect(doc_.leftMargin, h-(12*mm), doc_.width, 3, fill=1, stroke=0)

        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.HexColor('#777'))
        canvas.drawString(doc_.leftMargin, 9*mm, f"Relatório — {mes:02d}/{ano} — IJASystem")
        canvas.drawRightString(doc_.leftMargin + doc_.width, 9*mm, f"Página {canvas.getPageNumber()}")
        canvas.restoreState()

    doc.build(story, onFirstPage=_header_footer, onLaterPages=_header_footer)

    nome_arquivo = f"relatorio_IJASystem_{ano}_{mes:02d}"
    if uvis_id:
        nome_arquivo += f"_UVIS_{uvis_id}"

    return send_file(
        caminho_pdf,
        as_attachment=True,
        download_name=f"{nome_arquivo}.pdf",
        mimetype="application/pdf"
    )


# =======================================================================
# ROTA 3: Exportar Excel (Com Filtro UVIS)
# =======================================================================
@bp.route('/admin/exportar_relatorio_excel')
@login_required
def exportar_relatorio_excel():
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # -------------------------
    # 1. Parâmetros e filtros
    # -------------------------
    mes = request.args.get('mes', datetime.now().month, type=int)
    ano = request.args.get('ano', datetime.now().year, type=int)
    orient = request.args.get('orient', default='portrait')  # caso queira extensão futura
    filtro_data = f"{ano}-{mes:02d}"

    # Controle de acesso UVIS
    if current_user.tipo_usuario == 'uvis':
        uvis_id = current_user.id
    else:
        uvis_id = request.args.get('uvis_id', type=int)

    # -------------------------
    # 2. Busca de Dados
    # -------------------------
    query_dados = db.session.query(
        Solicitacao.id,
        Solicitacao.status,
        Solicitacao.foco,
        Solicitacao.tipo_visita,
        Solicitacao.altura_voo,
        Solicitacao.data_agendamento,
        Solicitacao.hora_agendamento,
        Solicitacao.cep,
        Solicitacao.logradouro,
        Solicitacao.numero,
        Solicitacao.bairro,
        Solicitacao.cidade,
        Solicitacao.uf,
        Solicitacao.latitude,
        Solicitacao.longitude,
        Usuario.nome_uvis,
        Usuario.regiao
    ).join(Usuario, Usuario.id == Solicitacao.usuario_id)

    # Filtro de data compatível com PostgreSQL/SQLite
    if db.engine.name == 'postgresql':
        query_dados = query_dados.filter(db.func.to_char(Solicitacao.data_criacao, 'YYYY-MM') == filtro_data)
    else:
        query_dados = query_dados.filter(db.func.strftime('%Y-%m', Solicitacao.data_criacao) == filtro_data)

    # Filtro opcional por UVIS
    if uvis_id:
        query_dados = query_dados.filter(Solicitacao.usuario_id == uvis_id)

    dados = query_dados.order_by(Solicitacao.data_criacao.desc()).all()

    # -------------------------
    # 3. Criar arquivo Excel
    # -------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório"

    colunas = [
        "ID", "Status", "Foco", "Tipo Visita", "Altura Voo",
        "Data Agendamento", "Hora Agendamento",
        "CEP", "Logradouro", "Número", "Bairro", "Cidade", "UF",
        "Latitude", "Longitude", "UVIS", "Região"
    ]

    header_fill = PatternFill(start_color="1E90FF", end_color="1E90FF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style='thin', color="000000")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    zebra1 = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
    zebra2 = PatternFill(start_color="FFF7FBFF", end_color="FFF7FBFF", fill_type="solid")

    # Cabeçalho
    for col_num, col_name in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=col_num, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    # Preenchimento de linhas
    for row_num, row in enumerate(dados, 2):
        data_agendamento_fmt = row.data_agendamento.strftime("%d/%m/%Y") if row.data_agendamento else ""
        hora_agendamento_fmt = row.hora_agendamento.strftime("%H:%M") if row.hora_agendamento else ""

        values = [
            row.id,
            row.status,
            row.foco,
            row.tipo_visita,
            row.altura_voo,
            data_agendamento_fmt,
            hora_agendamento_fmt,
            row.cep,
            row.logradouro,
            row.numero,
            row.bairro,
            row.cidade,
            row.uf,
            row.latitude,
            row.longitude,
            row.nome_uvis,
            row.regiao
        ]

        for col_index, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_index, value=value)
            cell.border = thin_border
            if col_index in (1, 3, 6, 8, 15, 16):
                cell.alignment = center
            else:
                cell.alignment = Alignment(vertical="top", horizontal="left")
            cell.fill = zebra1 if (row_num % 2 == 0) else zebra2

    # Ajuste de largura das colunas
    for col in ws.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 60)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(colunas))}1"

    # Gerar arquivo em memória
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Nome do arquivo
    nome_arquivo = f"relatorio_IJASystem_{ano}_{mes:02d}"
    if uvis_id:
        nome_arquivo += f"_UVIS_{uvis_id}"

    return send_file(
        output,
        download_name=f"{nome_arquivo}.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )



from flask import flash, redirect, url_for, render_template, request
from flask_login import login_required, current_user
from app import db
from app.models import Solicitacao, Usuario
from datetime import datetime
from sqlalchemy.orm import joinedload

@bp.route('/admin/editar_completo/<int:id>', methods=['GET', 'POST'])
@login_required
def admin_editar_completo(id):
    # 🔐 Controle de acesso
    if current_user.tipo_usuario != 'admin':
        flash('Permissão negada. Apenas administradores podem acessar esta página.', 'danger')
        return redirect(url_for('main.admin_dashboard'))

    # Busca segura com joinedload para evitar lazy-loading
    pedido = Solicitacao.query.options(joinedload(Solicitacao.usuario)).get_or_404(id)

    # Listas para selects do template (pré-preenchimento)
    status_opcoes = ["PENDENTE", "EM ANÁLISE", "APROVADO", "APROVADO COM RECOMENDAÇÕES", "NEGADO"]
    foco_opcoes = ["Foco 1", "Foco 2", "Foco 3"]  # ajuste conforme seus valores reais
    tipo_visita_opcoes = ["Tipo 1", "Tipo 2", "Tipo 3"]  # ajuste conforme seus valores reais
    uf_opcoes = ["AC","AL","AP","AM","BA","CE","DF","ES","GO","MA","MT","MS","MG",
                 "PA","PB","PR","PE","PI","RJ","RN","RS","RO","RR","SC","SP","SE","TO"]

    if request.method == 'POST':
        try:
            # Guardar estado anterior de data/hora
            antes_data = pedido.data_agendamento
            antes_hora = pedido.hora_agendamento

            # 1️⃣ Atualizar datas e horas
            data_str = request.form.get('data_agendamento')
            hora_str = request.form.get('hora_agendamento')

            pedido.data_agendamento = datetime.strptime(data_str, '%Y-%m-%d').date() if data_str else None
            pedido.hora_agendamento = datetime.strptime(hora_str, '%H:%M').time() if hora_str else None

            # 2️⃣ Atualizar campos principais
            pedido.foco = request.form.get('foco') or pedido.foco
            pedido.tipo_visita = request.form.get('tipo_visita') or pedido.tipo_visita
            pedido.altura_voo = request.form.get('altura_voo') or pedido.altura_voo
            pedido.apoio_cet = request.form.get('apoio_cet', 'não').lower() == 'sim'
            pedido.observacao = request.form.get('observacao') or pedido.observacao

            # 3️⃣ Atualizar endereço
            pedido.cep = request.form.get('cep') or pedido.cep
            pedido.logradouro = request.form.get('logradouro') or pedido.logradouro
            pedido.numero = request.form.get('numero') or pedido.numero
            pedido.bairro = request.form.get('bairro') or pedido.bairro
            pedido.cidade = request.form.get('cidade') or pedido.cidade
            pedido.uf = request.form.get('uf') or pedido.uf
            pedido.complemento = request.form.get('complemento') or pedido.complemento

            # 4️⃣ Atualizar protocolo, status, justificativa e coordenadas
            pedido.protocolo = request.form.get('protocolo') or pedido.protocolo
            pedido.status = request.form.get('status') or pedido.status
            pedido.justificativa = request.form.get('justificativa') or pedido.justificativa

            lat = request.form.get('latitude')
            lon = request.form.get('longitude')
            pedido.latitude = float(lat) if lat else None
            pedido.longitude = float(lon) if lon else None

            # Commit
            db.session.commit()

            # 🔔 Notificação se agendamento mudou
            mudou_agendamento = (antes_data != pedido.data_agendamento) or (antes_hora != pedido.hora_agendamento)
            if pedido.data_agendamento and mudou_agendamento:
                data_fmt = pedido.data_agendamento.strftime("%d/%m/%Y")
                hora_fmt = pedido.hora_agendamento.strftime("%H:%M") if pedido.hora_agendamento else "00:00"
                criar_notificacao(
                    usuario_id=pedido.usuario_id,
                    titulo="Agendamento atualizado",
                    mensagem=f"Sua solicitação foi agendada para {data_fmt} às {hora_fmt}.",
                    link=url_for("main.agenda")
                )

            flash('Solicitação atualizada com sucesso!', 'success')
            return redirect(url_for('main.admin_dashboard'))

        except ValueError as ve:
            db.session.rollback()
            flash(f"Erro no formato de data/hora: {ve}", 'warning')
        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao salvar a solicitação: {e}", 'danger')

    return render_template(
        'admin_editar_completo.html',
        pedido=pedido,
        status_opcoes=status_opcoes,
        foco_opcoes=foco_opcoes,
        tipo_visita_opcoes=tipo_visita_opcoes,
        uf_opcoes=uf_opcoes
    )

from flask_login import current_user, login_required

@bp.route('/admin/deletar/<int:id>', methods=['POST'], endpoint='deletar_registro')
@login_required
def deletar(id):
    # Verifica se é admin
    if current_user.tipo_usuario != 'admin':  # <-- CORRETO: tipo_usuario
        flash('Permissão negada. Apenas administradores podem deletar registros.', 'danger')
        return redirect(url_for('main.admin_dashboard'))

    # Busca a solicitação
    pedido = Solicitacao.query.get_or_404(id)
    pedido_id = pedido.id

    # Nome do autor da solicitação
    autor_nome = pedido.usuario.nome_uvis if pedido.usuario else "UVIS"

    try:
        db.session.delete(pedido)
        db.session.commit()
    except Exception:
        db.session.rollback()
        # Não mostra erro ao usuário

    flash(f"Pedido #{pedido_id} da {autor_nome} deletado permanentemente.", "success")
    return redirect(url_for('main.admin_dashboard'))


from flask_login import login_required, current_user

from flask_login import login_required, current_user
import traceback

from flask import request, render_template
from flask_login import login_required, current_user
from sqlalchemy.orm import joinedload
from datetime import datetime
import json

@bp.route("/agenda")
@login_required
def agenda():
    try:
        # --- Usuário atual ---
        user_tipo = current_user.tipo_usuario
        user_id = current_user.id

        # --- Filtros GET ---
        filtro_status = request.args.get("status") or None
        filtro_uvis_id = request.args.get("uvis_id", type=int)
        mes = request.args.get("mes", datetime.now().month, type=int)
        ano = request.args.get("ano", datetime.now().year, type=int)
        d = request.args.get("d")
        initial_date = d or f"{ano}-{mes:02d}-01"

        # --- Query base ---
        query = Solicitacao.query.options(joinedload(Solicitacao.usuario))

        if user_tipo not in ["admin", "operario", "visualizar"]:
            query = query.filter(Solicitacao.usuario_id == user_id)
            filtro_uvis_id = None
            pode_filtrar_uvis = False
        else:
            pode_filtrar_uvis = True
            if filtro_uvis_id:
                query = query.filter(Solicitacao.usuario_id == filtro_uvis_id)

        if filtro_status:
            query = query.filter(Solicitacao.status == filtro_status)

        filtro_mesano = f"{ano}-{mes:02d}"
        if db.engine.name == 'postgresql':
            query = query.filter(db.func.to_char(Solicitacao.data_agendamento, "YYYY-MM") == filtro_mesano)
        else:
            query = query.filter(db.func.strftime("%Y-%m", Solicitacao.data_agendamento) == filtro_mesano)

        eventos = query.all()

        # --- Monta eventos para o FullCalendar ---
        agenda_eventos = []
        for e in eventos:
            try:
                data = e.data_agendamento.strftime("%Y-%m-%d")
                hora = e.hora_agendamento.strftime("%H:%M") if e.hora_agendamento else "00:00"
                uvis_nome = e.usuario.nome_uvis if e.usuario else "UVIS"
            except:
                uvis_nome = "UVIS"

            ev = {
                "id": str(e.id),
                "title": f"{e.foco} - {uvis_nome}",
                "start": f"{data}T{hora}",
                "color": (
                    "#198754" if e.status == "APROVADO" else
                    "#ffa023" if e.status == "APROVADO COM RECOMENDAÇÕES" else
                    "#dc3545" if e.status == "NEGADO" else
                    "#e9fa05" if e.status == "EM ANÁLISE" else
                    "#0d6efd"
                ),
                "extendedProps": {
                    "foco": e.foco,
                    "uvis": uvis_nome,
                    "hora": hora,
                    "status": e.status
                }
            }
            agenda_eventos.append(ev)

        # --- Variáveis para filtros ---
        status_opcoes = ["PENDENTE", "EM ANÁLISE", "APROVADO", "APROVADO COM RECOMENDAÇÕES", "NEGADO"]

        uvis_disponiveis = []
        if user_tipo in ["admin", "operario", "visualizar"]:
            uvis_disponiveis = db.session.query(Usuario.id, Usuario.nome_uvis).filter(Usuario.tipo_usuario == "uvis").order_by(Usuario.nome_uvis).all()

        # --- Anos disponíveis ---
        if db.engine.name == 'postgresql':
            func_ano = db.func.to_char(Solicitacao.data_agendamento, "YYYY")
        else:
            func_ano = db.func.strftime("%Y", Solicitacao.data_agendamento)

        anos_raw = db.session.query(func_ano).filter(Solicitacao.data_agendamento.isnot(None)).distinct().order_by(func_ano.desc()).all()
        anos_disponiveis = [int(a[0]) for a in anos_raw if a and a[0]]
        if not anos_disponiveis:
            anos_disponiveis = [datetime.now().year]

        # --- Dicionário de filtros para template ---
        filtros = {
            "uvis_id": filtro_uvis_id,
            "status": filtro_status,
            "mes": mes,
            "ano": ano
        }

        return render_template(
            "agenda.html",
            eventos_json=json.dumps(agenda_eventos),
            filtros=filtros,
            status_opcoes=status_opcoes,
            uvis_disponiveis=uvis_disponiveis,
            anos_disponiveis=anos_disponiveis,
            initial_date=initial_date,
            pode_filtrar_uvis=pode_filtrar_uvis
        )

    except Exception as e:
        import traceback
        print("TRACEBACK COMPLETO:")
        traceback.print_exc()
        return f"ERRO NA AGENDA: {str(e)}"


@bp.route("/agenda/exportar_excel", endpoint="agenda_exportar_excel")
@login_required
def exportar_excel():  # <--- função interna com nome diferente
    if current_user.tipo_usuario != "admin":
        abort(403)  # Forbidden

    user_tipo = current_user.tipo_usuario
    user_id = current_user.id
    export_all = request.args.get("all") == "1"

    # filtros
    filtro_status = None if export_all else (request.args.get("status") or None)
    filtro_uvis_id = None if export_all else request.args.get("uvis_id", type=int)
    mes = None if export_all else request.args.get("mes", type=int)
    ano = None if export_all else request.args.get("ano", type=int)

    query = Solicitacao.query.options(joinedload(Solicitacao.usuario))

    if filtro_uvis_id:
        query = query.filter(Solicitacao.usuario_id == filtro_uvis_id)
    if filtro_status:
        query = query.filter(Solicitacao.status == filtro_status)
    if mes and ano:
        filtro_mesano = f"{ano}-{mes:02d}"
        if db.engine.name == 'postgresql':
            query = query.filter(db.func.to_char(Solicitacao.data_agendamento, "YYYY-MM") == filtro_mesano)
        else:
            query = query.filter(db.func.strftime("%Y-%m", Solicitacao.data_agendamento) == filtro_mesano)

    query = query.order_by(
        Solicitacao.data_agendamento.desc(),
        Solicitacao.hora_agendamento.desc()
    )
    eventos = query.all()
    # -----------------------------
    # Monta XLSX
    # -----------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Agenda"

    headers = [
        "DATA",
        "HORÁRIO",
        "REGIÃO",
        "UVIS",
        "CET",
        "ENDEREÇO DA AÇÃO",
        "CEP",
        "FOCO DA AÇÃO",
        "COORDENADA GEOGRÁFICA",
        "Altura dos Voos",
        "Protocolo DECA",
        "Status",
    ]
    ws.append(headers)

    for p in eventos:
        endereco_completo = (
            f"{p.logradouro or ''}, {getattr(p, 'numero', '')} - "
            f"{p.bairro or ''} - "
            f"{(p.cidade or '')}/{(p.uf or '')} - "
            f"{p.cep or ''}"
        )
        if getattr(p, "complemento", None):
            endereco_completo += f" - {p.complemento}"

        cet_txt = "SIM" if getattr(p, "apoio_cet", None) else "NÃO"
        data_str = p.data_agendamento.strftime("%d/%m/%Y") if p.data_agendamento else ""
        hora_str = p.hora_agendamento.strftime("%H:%M") if p.hora_agendamento else ""
        uvis_nome = p.usuario.nome_uvis if getattr(p, "usuario", None) else ""
        regiao = p.usuario.regiao if getattr(p, "usuario", None) else ""
        lat = getattr(p, "latitude", "") or ""
        lon = getattr(p, "longitude", "") or ""
        coordenada = f"{lat},{lon}" if (lat or lon) else ""
        protocolo_deca = getattr(p, "protocolo_deca", None) or getattr(p, "protocolo", "") or ""

        ws.append([
            data_str,
            hora_str,
            regiao,
            uvis_nome,
            cet_txt,
            endereco_completo,
            getattr(p, "cep", "") or "",
            getattr(p, "foco", "") or "",
            coordenada,
            getattr(p, "altura_voo", "") or "",
            protocolo_deca,
            getattr(p, "status", "") or "",
        ])

    # -----------------------------
    # Estilo
    # -----------------------------
    header_fill = PatternFill("solid", fgColor="0D6EFD")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap = Alignment(vertical="top", wrap_text=True)

    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center

    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border
            cell.alignment = wrap if cell.row > 1 else center

    for col in range(1, ws.max_column + 1):
        max_len = max(len(str(c.value)) if c.value else 0 for c in ws[get_column_letter(col)])
        ws.column_dimensions[get_column_letter(col)].width = min(max(12, max_len + 2), 60)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    nome = "agenda_tudo.xlsx" if export_all else "agenda_exportada.xlsx"

    return send_file(
        bio,
        as_attachment=True,
        download_name=nome,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
# =================================================
# NOTIFICAÇÕES (Flask-Login: login_required + current_user)
# Requer no topo:
# from flask_login import login_required, current_user
# from flask import abort, redirect, url_for, render_template
# from datetime import datetime, date
# from sqlalchemy.orm import joinedload
# =================================================

# -------------------------------------------------
# CRIAR NOTIFICAÇÃO
# -------------------------------------------------
def criar_notificacao(usuario_id, titulo, mensagem="", link=None):
    n = Notificacao(
        usuario_id=usuario_id,
        titulo=titulo,
        mensagem=mensagem or "",
        link=link
    )
    db.session.add(n)
    db.session.commit()
    return n


# -------------------------------------------------
# GARANTIR NOTIFICAÇÕES DO DIA (sem duplicar)
# - UVIS: cria apenas para ela mesma
# Observação: com soft delete (apagada_em), NÃO recria se já existiu (mesmo apagada),
# porque ela continua existindo no banco.
# -------------------------------------------------
def garantir_notificacoes_do_dia(usuario_id):
    hoje = date.today()

    ags = (
        Solicitacao.query
        .options(joinedload(Solicitacao.usuario))
        .filter_by(usuario_id=usuario_id)
        .filter(Solicitacao.data_agendamento == hoje)
        .all()
    )

    for s in ags:
        hora_fmt = s.hora_agendamento.strftime("%H:%M") if s.hora_agendamento else "00:00"

        # 🔒 chave estável
        link = url_for("main.agenda", sid=s.id, d=hoje.isoformat())

        ja_existe = (
            Notificacao.query
            .filter_by(usuario_id=usuario_id, link=link)
            .first()
        )
        if ja_existe:
            continue

        criar_notificacao(
            usuario_id=usuario_id,
            titulo="Agendamento para hoje",
            mensagem=f"Você tem um agendamento hoje às {hora_fmt} (Foco: {s.foco}).",
            link=link
        )


# -------------------------------------------------
# LER NOTIFICAÇÃO
# -------------------------------------------------
@bp.route("/notificacoes/<int:notif_id>/ler")
@login_required
def ler_notificacao(notif_id):
    user_tipo = current_user.tipo_usuario

    if user_tipo in ["admin", "operario", "visualizar"]:
        n = Notificacao.query.get_or_404(notif_id)
    else:
        n = (Notificacao.query
             .filter_by(id=notif_id, usuario_id=current_user.id)
             .first_or_404())

    if n.lida_em is None:
        n.lida_em = datetime.utcnow()
        db.session.commit()

    return redirect(n.link or url_for("main.notificacoes"))


# -------------------------------------------------
# LISTAR NOTIFICAÇÕES
# -------------------------------------------------
@bp.route("/notificacoes")
@login_required
def notificacoes():
    user_tipo = current_user.tipo_usuario

    # ✅ só UVIS gera lembrete do dia (pro próprio usuário)
    if user_tipo not in ["admin", "operario", "visualizar"]:
        garantir_notificacoes_do_dia(current_user.id)

    base = Notificacao.query.filter(Notificacao.apagada_em.is_(None))

    # ✅ admin/operário/visualizar vê tudo, uvis só as dela
    if user_tipo in ["admin", "operario", "visualizar"]:
        itens = base.order_by(Notificacao.criada_em.desc()).all()
    else:
        itens = (base
                 .filter_by(usuario_id=current_user.id)
                 .order_by(Notificacao.criada_em.desc())
                 .all())

    return render_template("notificacoes.html", itens=itens)


# -------------------------------------------------
# EXCLUIR UMA NOTIFICAÇÃO (SOFT DELETE)
# -------------------------------------------------
@bp.route("/notificacoes/<int:notif_id>/excluir", methods=["POST"])
@login_required
def excluir_notificacao(notif_id):
    user_tipo = current_user.tipo_usuario

    if user_tipo in ["admin", "operario", "visualizar"]:
        n = Notificacao.query.get_or_404(notif_id)
    else:
        n = (Notificacao.query
             .filter_by(id=notif_id, usuario_id=current_user.id)
             .first_or_404())

    n.apagada_em = datetime.utcnow()
    db.session.commit()

    return redirect(url_for("main.notificacoes"))


# -------------------------------------------------
# LIMPAR TODAS AS NOTIFICAÇÕES (SOFT DELETE EM LOTE)
# -------------------------------------------------
@bp.route("/notificacoes/limpar", methods=["POST"])
@login_required
def limpar_notificacoes():
    user_tipo = current_user.tipo_usuario
    agora = datetime.utcnow()

    q = Notificacao.query.filter(Notificacao.apagada_em.is_(None))

    if user_tipo not in ["admin", "operario", "visualizar"]:
        q = q.filter_by(usuario_id=current_user.id)

    q.update({"apagada_em": agora}, synchronize_session=False)
    db.session.commit()

    return redirect(url_for("main.notificacoes"))

# ==========================
# CHATBOT UVIS (FAQ inteligente)
# ==========================
import re
import unicodedata

from flask import jsonify, request
from flask_login import login_required, current_user


def _norm(text: str) -> str:
    if not text:
        return ""
    text = text.strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"\s+", " ", text)
    return text


UVIS_FAQ = [
    {
        "title": "Status da solicitação",
        "keywords": ["status", "pendente", "em analise", "aprovado", "negado", "protocolo"],
        "answer": (
            "📌 **Significado dos status**:\n"
            "- **Pendente**: solicitação registrada e aguardando início do processo.\n"
            "- **Em Análise**: pedido em validação pela equipe responsável.\n"
            "- **Aprovado**: pedido autorizado (pode aparecer o número de protocolo).\n"
            "- **Aprovado com Recomendações**: pedido aprovado com sugestões de melhoria.\n"
            "- **Negado**: pedido não aprovado (o motivo aparece nos detalhes).\n\n"
            "💡 Dica: clique em **Detalhes** para ver justificativa/protocolo."
        ),
    },
    {
        "title": "O que tem na tela 'Minhas Solicitações' (Dashboard)",
        "keywords": ["dashboard", "minhas solicitacoes", "tela inicial", "filtro", "detalhes", "nova solicitacao"],
        "answer": (
            "Na tela **Minhas Solicitações** você encontra:\n"
            "- Botão **Nova Solicitação** (abre o formulário)\n"
            "- **Filtro por status** (Pendente, Em Análise, Aprovado, Aprovado com Recomendações, Negado)\n"
            "- **Tabela** com data/hora, localização e foco\n"
            "- Botão **Detalhes** (abre um modal com informações completas)\n"
        ),
    },
    {
        "title": "Campos obrigatórios ao criar uma solicitação",
        "keywords": ["novo", "nova solicitacao", "cadastro", "campos", "obrigatorio", "cep", "numero", "tipo de visita", "altura", "foco"],
        "answer": (
            "✅ No cadastro de uma nova solicitação, atenção aos campos:\n"
            "- **Data** e **Hora** (obrigatórios)\n"
            "- **CEP** (8 dígitos) para preencher endereço automático\n"
            "- **Logradouro** (confirmar) e **Número** (preencher manualmente)\n"
            "- **Tipo de visita** (Monitoramento / Aedes / Culex)\n"
            "- **Altura do voo** (10m, 20m, 30m, 40m)\n"
            "- **Foco da ação** (ex.: Imóvel Abandonado, Piscina/Caixa d’água, Terreno Baldio, Ponto Estratégico)\n"
        ),
    },
    {
        "title": "CEP / endereço não encontrado e boas práticas",
        "keywords": ["cep", "endereco", "logradouro", "bairro", "cidade", "uf", "nao encontrado", "boas praticas"],
        "answer": (
            "Se o **CEP não for encontrado**, preencha o endereço manualmente e revise.\n"
            "Boas práticas:\n"
            "- confira se o **CEP** corresponde ao local\n"
            "- verifique logradouro/bairro/cidade/UF\n"
            "- preencha o **número** (sem ele pode dificultar a localização)\n"
        ),
    },
    {
        "title": "Latitude/Longitude e mapa",
        "keywords": ["latitude", "longitude", "coordenadas", "gps", "mapa"],
        "answer": (
            "📍 **Latitude/Longitude** é opcional (recomendado) e melhora a precisão.\n"
            "Se houver coordenadas, o sistema pode oferecer acesso rápido ao mapa."
        ),
    },
    {
        "title": "Notificações e Agenda",
        "keywords": ["notificacao", "notificacoes", "agenda", "calendario", "lembrete"],
        "answer": (
            "🔔 Em **Notificações**, você vê alertas da unidade (lembretes do dia/atualizações).\n"
            "Ao clicar, pode ser direcionado para a **Agenda**, que mostra os agendamentos por mês/semana/lista."
        ),
    },
    {
        "title": "Checklist antes de enviar",
        "keywords": ["checklist", "antes de enviar", "enviar pedido", "validar"],
        "answer": (
            "🧾 **Checklist rápido antes de enviar**:\n"
            "☐ Data e hora corretas\n"
            "☐ CEP válido e endereço conferido\n"
            "☐ Número preenchido\n"
            "☐ Tipo de visita e altura do voo selecionados\n"
            "☐ Foco da ação selecionado\n"
            "☐ Observações (se necessário) com informações objetivas\n"
        ),
    },
    {
        "title": "Suporte",
        "keywords": ["suporte", "erro", "acesso", "login", "senha"],
        "answer": (
            "Se a dúvida for de **erro de acesso**, inconsistência de **CEP/endereço**, ou algo fora do fluxo: "
            "entre em contato com o time de desenvolvimento/suporte da IJA."
        ),
    },
]


@bp.route("/api/uvis/chatbot", methods=["POST"])
@login_required
def uvis_chatbot():
    # (opcional) se quiser limitar só para UVIS:
    # if current_user.tipo_usuario != "uvis":
    #     return jsonify({"answer": "Acesso negado."}), 403

    payload = request.get_json(silent=True) or {}
    msg = (payload.get("message") or "").strip()

    if not msg:
        return jsonify({"answer": "Escreva sua dúvida (ex.: “o que significa Em Análise?”)."}), 400

    nmsg = _norm(msg)

    best = None
    best_score = 0

    for item in UVIS_FAQ:
        score = 0
        for kw in item["keywords"]:
            if kw in nmsg:
                score += 1
        if score > best_score:
            best_score = score
            best = item

    if not best or best_score == 0:
        sugestoes = [
            "• “O que significa Pendente/Em Análise/Aprovado/Aprovado com Recomendações/Negado?”",
            "• “Quais campos são obrigatórios na Nova Solicitação?”",
            "• “O que fazer se o CEP não encontrar?”",
            "• “Qual o checklist antes de enviar?”",
            "• “Como funciona Notificações e Agenda?”",
        ]
        return jsonify({
            "answer": (
                "Não encontrei essa dúvida diretamente no manual.\n\n"
                "Tenta uma dessas perguntas:\n" + "\n".join(sugestoes)
            ),
            "matched": None,
            "confidence": 0,
        }), 200

    return jsonify({
        "answer": best["answer"],
        "matched": best["title"],
        "confidence": best_score,
    }), 200


import os
from flask import abort, send_from_directory
from flask_login import login_required, current_user

@bp.route("/solicitacao/<int:id>/anexo", endpoint="baixar_anexo")
@bp.route("/admin/solicitacao/<int:id>/anexo", endpoint="baixar_anexo_admin")
@login_required
def baixar_anexo(id):
    pedido = Solicitacao.query.get_or_404(id)

    # 🔐 permissões
    if current_user.tipo_usuario not in ["admin", "operario", "visualizar", "uvis"]:
        abort(403)
    if current_user.tipo_usuario == "uvis" and pedido.usuario_id != current_user.id:
        abort(403)

    if not pedido.anexo_path:
        abort(404)

    # ✅ mesma pasta do upload
    upload_folder = get_upload_folder()

    # ✅ normaliza o caminho salvo no banco
    rel = (pedido.anexo_path or "").replace("\\", "/")
    if rel.startswith("upload-files/"):
        rel = rel.split("upload-files/", 1)[1]
    rel = os.path.basename(rel)  # segurança

    file_path = os.path.join(upload_folder, rel)
    if not os.path.isfile(file_path):
        abort(404)

    return send_from_directory(
        upload_folder,
        rel,
        as_attachment=False,
        download_name=(pedido.anexo_nome or rel)
    )


@bp.route("/admin/uvis/novo", methods=["GET", "POST"], endpoint="admin_uvis_novo")
@login_required
def admin_uvis_novo():
    # SOMENTE ADMIN
    if current_user.tipo_usuario != "admin":
        abort(403)

    if request.method == "POST":
        nome_uvis = (request.form.get("nome_uvis") or "").strip()
        regiao = (request.form.get("regiao") or "").strip() or None
        codigo_setor = (request.form.get("codigo_setor") or "").strip() or None

        login = (request.form.get("login") or "").strip()
        senha = request.form.get("senha") or ""
        confirmar = request.form.get("confirmar") or ""

        if not nome_uvis or not login or not senha:
            flash("Preencha: Nome da UVIS, Login e Senha.", "warning")
            return render_template("admin_uvis_novo.html")

        if senha != confirmar:
            flash("As senhas não conferem.", "warning")
            return render_template("admin_uvis_novo.html")

        novo_user = Usuario(
            nome_uvis=nome_uvis,
            regiao=regiao,
            codigo_setor=codigo_setor,
            login=login,
            tipo_usuario="uvis",
        )
        novo_user.set_senha(senha)

        try:
            db.session.add(novo_user)
            db.session.commit()
            flash("UVIS cadastrada com sucesso!", "success")
            return redirect(url_for("main.admin_dashboard"))
        except IntegrityError:
            db.session.rollback()
            flash("Esse login já está em uso. Escolha outro.", "danger")
        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao cadastrar UVIS: {e}", "danger")

    return render_template("admin_uvis_novo.html")


@bp.route("/admin/uvis", methods=["GET"], endpoint="admin_uvis_listar")
@login_required
def admin_uvis_listar():
    # SOMENTE ADMIN
    if current_user.tipo_usuario != "admin":
        abort(403)

    q = (request.args.get("q") or "").strip()
    regiao = (request.args.get("regiao") or "").strip()
    codigo_setor = (request.args.get("codigo_setor") or "").strip()

    query = Usuario.query.filter(Usuario.tipo_usuario == "uvis")

    if q:
        query = query.filter(
            db.or_(
                Usuario.nome_uvis.ilike(f"%{q}%"),
                Usuario.login.ilike(f"%{q}%")
            )
        )

    if regiao:
        query = query.filter(Usuario.regiao.ilike(f"%{regiao}%"))

    if codigo_setor:
        query = query.filter(Usuario.codigo_setor.ilike(f"%{codigo_setor}%"))

    page = request.args.get("page", 1, type=int)
    paginacao = query.order_by(Usuario.nome_uvis.asc()).paginate(
        page=page, per_page=10, error_out=False
    )

    return render_template(
        "admin_uvis_listar.html",
        uvis=paginacao.items,
        paginacao=paginacao,
        q=q,
        regiao=regiao,
        codigo_setor=codigo_setor
    )


@bp.route("/admin/uvis/<int:id>/editar", methods=["GET", "POST"], endpoint="admin_uvis_editar")
@login_required
def admin_uvis_editar(id):
    if current_user.tipo_usuario != "admin":
        abort(403)

    uvis = Usuario.query.get_or_404(id)

    if uvis.tipo_usuario != "uvis":
        flash("Registro inválido para edição.", "danger")
        return redirect(url_for("main.admin_uvis_listar"))

    if request.method == "POST":
        nome_uvis = (request.form.get("nome_uvis") or "").strip()
        regiao = (request.form.get("regiao") or "").strip() or None
        codigo_setor = (request.form.get("codigo_setor") or "").strip() or None
        login = (request.form.get("login") or "").strip()

        senha = (request.form.get("senha") or "").strip()
        confirmar = (request.form.get("confirmar") or "").strip()

        if not nome_uvis or not login:
            flash("Preencha: Nome da UVIS e Login.", "warning")
            return render_template("admin_uvis_editar.html", uvis=uvis)

        if senha:
            if senha != confirmar:
                flash("As senhas não conferem.", "warning")
                return render_template("admin_uvis_editar.html", uvis=uvis)
            uvis.set_senha(senha)

        uvis.nome_uvis = nome_uvis
        uvis.regiao = regiao
        uvis.codigo_setor = codigo_setor
        uvis.login = login

        try:
            db.session.commit()
            flash("UVIS atualizada com sucesso!", "success")
            return redirect(url_for("main.admin_uvis_listar"))
        except IntegrityError:
            db.session.rollback()
            flash("Esse login já está em uso. Escolha outro.", "danger")
        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao salvar: {e}", "danger")

    return render_template("admin_uvis_editar.html", uvis=uvis)


@bp.route("/admin/uvis/<int:id>/excluir", methods=["POST"], endpoint="admin_uvis_excluir")
@login_required
def admin_uvis_excluir(id):
    if current_user.tipo_usuario != "admin":
        abort(403)

    uvis = Usuario.query.get_or_404(id)

    if uvis.tipo_usuario != "uvis":
        flash("Registro inválido para exclusão.", "danger")
        return redirect(url_for("main.admin_uvis_listar"))

    existe = Solicitacao.query.filter_by(usuario_id=uvis.id).first()
    if existe:
        flash("Não é possível excluir: esta UVIS possui solicitações vinculadas.", "warning")
        return redirect(url_for("main.admin_uvis_listar"))

    try:
        db.session.delete(uvis)
        db.session.commit()
        flash("UVIS excluída com sucesso!", "success")
    except Exception:
        db.session.rollback()
        flash("Erro ao excluir UVIS.", "danger")

    return redirect(url_for("main.admin_uvis_listar"))

# ==========================
# CHATBOT ADMIN (FAQ inteligente) - Flask-Login
# ==========================
import re
import unicodedata

from flask import jsonify, request
from flask_login import login_required, current_user


def _norm_admin(text: str) -> str:
    if not text:
        return ""
    text = text.strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"\s+", " ", text)
    return text


def _clean_answer(text: str) -> str:
    """Remove markdown simples (**negrito**, `code`, etc) e normaliza."""
    if not text:
        return ""
    text = re.sub(r"\*\*(.*?)\*\*", r"\1", text)   # remove ** **
    text = text.replace("`", "")                  # remove ` `
    text = re.sub(r"\n{3,}", "\n\n", text)         # evita muitas quebras
    return text.strip()


ADMIN_FAQ = [
    {
        "title": "Perfis e permissões",
        "keywords": ["acesso", "perfil", "permissao", "permissões", "admin", "operario", "operário", "visualizar", "quem pode"],
        "answer": (
            "Perfis do painel:\n"
            "- Administrador: acesso total (editar, excluir, gerenciar UVIS, relatórios e agenda).\n"
            "- Operário: consegue salvar decisões (status/protocolo/justificativa).\n"
            "- Visualizar: apenas leitura.\n"
        ),
    },
    {
        "title": "Filtros no painel",
        "keywords": ["filtro", "filtrar", "status", "unidade", "uvis", "regiao", "região", "buscar", "pesquisar"],
        "answer": (
            "No painel você pode filtrar por:\n"
            "- Status\n"
            "- Unidade (UVIS)\n"
            "- Região\n"
            "Use os filtros para encontrar solicitações específicas rapidamente."),
    },
    {
        "title": "Olá! Como posso ajudar?",
        "keywords": ["olá", "oi", "hello", "hi", "bom dia", "boa tarde", "boa noite", "ajuda", "suporte"],
        "answer": (
            "Olá! Sou o assistente virtual do painel administrativo.\n"
            "Posso ajudar com dúvidas sobre:\n"
            "- Perfis e permissões\n"
            "- Filtros no painel\n"
            "- Salvar decisão\n"
            "- Editar completo\n"
            "- Excluir solicitação\n"
            "- Anexos\n"
            "- GPS e mapa\n"
            "- Exportar Excel do painel\n"
            "- Agenda\n"
            "- Relatórios\n"
            "- Gestão de UVIS\n"
            "Como posso ajudar você hoje?"
        ),
    },
    {
        "title": "Salvar decisão",
        "keywords": ["salvar", "decisao", "decisão", "status", "protocolo", "justificativa", "aprovado", "negado", "analise", "recomendacoes", "recomendações"],
        "answer": (
            "Em cada solicitação você pode definir:\n"
            "- Status\n"
            "- Protocolo\n"
            "- Justificativa (principalmente se negar ou orientar)\n"
            "Se o perfil for ‘Visualizar’, fica somente leitura."
        ),
    },
    {
        "title": "Editar completo",
        "keywords": ["editar", "editar completo", "corrigir", "alterar", "data", "hora", "endereco", "endereço", "agendamento"],
        "answer": (
            "Editar completo serve para corrigir todos os dados do pedido:\n"
            "data/hora, endereço, foco, tipo de visita, altura e observações.\n"
            "Em alguns casos o sistema pode gerar notificação para a unidade."
        ),
    },
    {
        "title": "Excluir solicitação",
        "keywords": ["excluir", "deletar", "apagar", "remover"],
        "answer": (
            "Excluir remove a solicitação definitivamente.\n"
            "Normalmente é restrito ao Administrador e pede confirmação."
        ),
    },
    {
        "title": "Anexos",
        "keywords": ["anexo", "arquivo", "upload", "baixar", "download", "pdf", "png", "jpg", "doc", "xlsx"],
        "answer": (
            "Você pode anexar arquivos na solicitação e depois baixar.\n"
            "Se o anexo não aparecer, verifique se foi salvo corretamente e se o arquivo é permitido."
        ),
    },
    {
        "title": "GPS e mapa",
        "keywords": ["gps", "latitude", "longitude", "coordenadas", "mapa", "google maps"],
        "answer": (
            "Latitude/Longitude ajudam na precisão.\n"
            "Quando preenchidas, o botão de mapa abre o local no Google Maps."
        ),
    },
    {
        "title": "Exportar Excel do painel",
        "keywords": ["exportar", "excel", "xlsx", "planilha", "baixar excel"],
        "answer": (
            "Existe exportação para Excel a partir do painel.\n"
            "Quando você usa filtros (status/unidade/região), isso tende a refletir no arquivo exportado."
        ),
    },
    {
        "title": "Agenda",
        "keywords": ["agenda", "calendario", "calendário", "eventos", "mes", "mês", "ano", "exportar agenda"],
        "answer": (
            "A Agenda mostra agendamentos por período.\n"
            "Você pode filtrar (quando disponível) e exportar."
        ),
    },
    {
        "title": "Relatórios",
        "keywords": ["relatorio", "relatórios", "pdf", "grafico", "gráfico", "totais", "mes", "ano"],
        "answer": (
            "Relatórios permitem filtrar por mês/ano e, quando disponível, por unidade.\n"
            "Também podem ter exportação em PDF e Excel."
        ),
    },
    {
        "title": "Gestão de UVIS",
        "keywords": ["uvis", "cadastrar uvis", "lista uvis", "gerenciar uvis", "unidade", "login", "senha", "codigo setor", "código setor", "regiao", "região"],
        "answer": (
            "Gestão de UVIS inclui:\n"
            "- Listar UVIS\n"
            "- Cadastrar UVIS\n"
            "- Editar UVIS (inclusive redefinir senha)\n"
            "Atenção: login não pode repetir."
        ),
    },
]


@bp.route("/api/admin/chatbot", methods=["POST"])
@login_required
def admin_chatbot():
    # 🔐 só perfis do painel
    if current_user.tipo_usuario not in ["admin", "operario", "visualizar"]:
        return jsonify({"answer": "Acesso negado para este chatbot."}), 403

    payload = request.get_json(silent=True) or {}
    msg = (payload.get("message") or "").strip()

    if not msg:
        return jsonify({"answer": "Digite sua dúvida (ex.: como exportar Excel?)."}), 400

    nmsg = _norm_admin(msg)

    best = None
    best_score = 0

    for item in ADMIN_FAQ:
        score = 0
        for kw in item["keywords"]:
            if kw in nmsg:
                score += 1
        if score > best_score:
            best_score = score
            best = item

    if not best or best_score == 0:
        sugestoes = [
            "Como filtrar por status/unidade/região?",
            "Como salvar decisão (status/protocolo/justificativa)?",
            "Como editar completo?",
            "Como exportar Excel?",
            "Como funciona Agenda/Relatórios?",
            "Como gerenciar UVIS?",
        ]
        return jsonify({
            "answer": "Não achei essa dúvida direto no guia.\n\nSugestões:\n- " + "\n- ".join(sugestoes),
            "matched": None,
            "confidence": 0,
        }), 200

    return jsonify({
        "answer": _clean_answer(best["answer"]),
        "matched": best["title"],
        "confidence": best_score,
    }), 200



@bp.app_errorhandler(404)
def pagina_nao_encontrada(e):
    return render_template(
        'erro.html', 
        codigo=404, 
        titulo="Página não encontrada", 
        mensagem="Ops! A página que você está procurando não existe ou foi movida."
    ), 404

@bp.app_errorhandler(500)
def erro_interno(e):
    # Opcional: printar o erro no terminal para você ver o que houve
    # print(f"Erro 500 detectado: {e}")
    return render_template(
        'erro.html', 
        codigo=500, 
        titulo="Erro Interno do Servidor", 
        mensagem="Desculpe, algo deu errado do nosso lado. Tente novamente mais tarde."
    ), 500
